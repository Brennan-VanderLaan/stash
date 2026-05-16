"""Bulk import of inventory from external tools.

Architecture
------------

This module is structured as a **per-source parser registry** with
a shared generic executor.  When a new product needs an importer
(Sortly, HomeBox, MyStuff2 Pro, …), add a parser function that
returns a :class:`ParseResult` and register it in :data:`PARSERS`.
The execution path — get-or-create Location → Floor → Rooms → loose
Boxes → Items — is product-agnostic, so a new source costs roughly
one header map + one parser function.

Currently registered:

* :data:`encircle` — Encircle's consumer "Detailed Spreadsheet"
  export (XLSX with embedded images plus a separate media ZIP).
  Encircle's consumer Home Inventory product shut down on
  2025-12-17 leaving a documented cohort of displaced users
  looking for a replacement.  V1 handles text only — photo
  extraction from ``xl/media/`` + the media ZIP companion ship
  in a follow-up commit.

Design notes
~~~~~~~~~~~~

* **One Location per import job.**  Auto-created with a
  source-tagged + timestamped name (e.g.
  ``"Imported from Encircle (2026-05-16 19:00)"``) so the
  operator can ``undo_import`` cleanly via a cascade-delete.
  No per-row provenance tracking needed.
* **Rooms map 1:1 from the source** to a Stash room under the
  import's auto-created floor.  Unique room names become
  unique rooms; duplicate items for the same room cohabit.
* **Loose-items box per room** via the existing
  :func:`dao.boxes.get_or_create_loose_for_room` helper.  The
  operator can later promote loose items into proper boxes
  through the /queue or /home flows.  This re-uses the loose-
  box feature we shipped for feedback #39.
* **Notes are the dumping ground** for source-side metadata
  Stash doesn't yet have first-class columns for (brand,
  model number, serial number, warranty …).  ``format_item_notes``
  centralises the formatting so search + audit reads stay
  consistent regardless of which source produced the row.

Edge cases the parsers defend against
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

* Header strings with stray whitespace, "#" vs "Number" vs
  "No", case variations — normalised before lookup.
* Numeric columns carrying the literal string ``"Receipt"`` (an
  Encircle quirk — price is on a paper receipt, not in the
  spreadsheet).  Skipped silently rather than coerced.
* ISO-8601 durations in warranty fields (``P7Y`` = 7 years,
  ``P6M`` = 6 months).  Humanised for the notes blob.
* UPC codes embedded in Notes (``Upc: NNNNNNN``).  Preserved
  verbatim in the merged notes blob.
* Entirely blank rows — silently skipped.
* Missing serial numbers — common; left empty.

Reference: ``github.com/tokendad/NesVentory`` (MIT, the only
purpose-built Encircle migrator) — consulted for the real export
schema and filename conventions.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable

import obs
from dao._base import (
    Actor,
    ForbiddenError,
    NotFoundError,
    db,
    require_role,
)
from dao import boxes as dao_boxes
from dao import items as dao_items


_log = obs.get_logger("dao.imports")


# Cap an import at 5,000 rows.  Plenty for a household inventory
# (typical Encircle exports run 100-800 rows; insurance-tier
# households top out around 2,000); guards against an accidentally
# uploaded billing CSV with a million rows.
MAX_IMPORT_ROWS = 5_000

# Cap upload file size at 10 MB.  Text-only XLSX exports sit
# under 2 MB even for large households; a 10 MB upload either has
# photos (V2 territory) or is malformed.
MAX_IMPORT_BYTES = 10 * 1024 * 1024


# ── Generic types ───────────────────────────────────────────────────


@dataclass
class ParseResult:
    """Output of any registered parser.

    ``items`` is a list of dicts keyed by internal field names (see
    :data:`KNOWN_INTERNAL_FIELDS`).  Parsers normalise their
    source-specific header strings into these keys so the executor
    doesn't care which product produced the rows.

    ``unmapped_headers`` lists header strings the parser didn't
    know how to bind — surfaced to the operator as a warning so
    they can decide whether to re-map columns and re-upload.
    """
    items: list[dict[str, str]] = field(default_factory=list)
    unmapped_headers: list[str] = field(default_factory=list)
    mapped_headers: list[str] = field(default_factory=list)
    total_rows: int = 0
    source: str = ""    # filled in by ``parse()`` at dispatch time


# Internal field names every parser should produce when possible.
# A parser that doesn't have a source-side equivalent for a given
# field simply leaves it out; the executor + notes formatter both
# tolerate sparse rows.
KNOWN_INTERNAL_FIELDS: set[str] = {
    "item_name", "brand", "model_number", "serial_number",
    "quantity", "purchase_vendor", "purchase_date", "purchase_price",
    "estimated_value", "warranty_duration",
    "extended_warranty_policy", "extended_warranty_phone",
    "notes", "room", "box", "disposition", "item_age",
    "pre_loss_condition", "upc",
}


# ── Generic value cleaners ──────────────────────────────────────────


def _clean_price(raw: Any) -> str:
    """Strip currency symbols + thousands commas; skip the literal
    string ``"Receipt"`` (an Encircle quirk).  Returns a string
    so we don't fight Python over int vs float for "$1,234.50"."""
    if raw is None or raw == "":
        return ""
    s = str(raw).strip()
    if s.lower() == "receipt":
        return ""
    s = re.sub(r"^[\$£€]\s*", "", s)
    s = s.replace(",", "")
    return s


_ISO_DURATION = re.compile(
    r"^P(?:(?P<y>\d+)Y)?(?:(?P<m>\d+)M)?(?:(?P<d>\d+)D)?$"
)


def _humanize_warranty(raw: Any) -> str:
    """Convert ISO-8601 duration strings (``P7Y`` = 7 years) to
    human-readable form.  Passes anything that doesn't match
    through verbatim so hand-typed warranty notes survive."""
    if raw is None or raw == "":
        return ""
    s = str(raw).strip()
    m = _ISO_DURATION.match(s)
    if not m:
        return s
    parts: list[str] = []
    if m.group("y"):
        n = int(m.group("y"))
        parts.append(f"{n} year{'' if n == 1 else 's'}")
    if m.group("m"):
        n = int(m.group("m"))
        parts.append(f"{n} month{'' if n == 1 else 's'}")
    if m.group("d"):
        n = int(m.group("d"))
        parts.append(f"{n} day{'' if n == 1 else 's'}")
    return " ".join(parts) if parts else s


def _clean_date(raw: Any) -> str:
    """Pass an ISO date through; coerce a datetime to its date
    portion (openpyxl auto-coerces XLSX date cells to datetimes)."""
    if raw is None or raw == "":
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if hasattr(raw, "isoformat"):
        return raw.isoformat()
    return str(raw).strip()


def _stringy(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def _normalise_header(s: str) -> str:
    """Header-name normalisation: lower-case, collapse whitespace,
    strip trailing punctuation that doesn't survive Excel ↔ CSV
    round-trips cleanly."""
    s = (s or "").lower().strip()
    s = re.sub(r"[‘’]", "'", s)
    s = re.sub(r"[“”]", '"', s)
    s = re.sub(r"\s+", " ", s)
    s = s.rstrip(":")
    return s


# ── Generic row mapper ──────────────────────────────────────────────


def _row_to_normalised(
    headers: list[str], row: list[Any], lookup: dict[str, str],
) -> dict[str, str]:
    """Map a raw row + its header row into a dict keyed by internal
    field names.  ``lookup`` is the per-source normalised-header
    → internal-field dict.  Unmapped columns drop on the floor
    (they're surfaced separately via :class:`ParseResult` so the
    UI can warn the operator without failing the whole import)."""
    out: dict[str, str] = {}
    for idx, header in enumerate(headers):
        internal = lookup.get(_normalise_header(header))
        if internal is None:
            continue
        if idx >= len(row):
            continue
        cell = row[idx]
        if internal in ("purchase_price", "estimated_value"):
            out[internal] = _clean_price(cell)
        elif internal == "warranty_duration":
            out[internal] = _humanize_warranty(cell)
        elif internal == "purchase_date":
            out[internal] = _clean_date(cell)
        else:
            out[internal] = _stringy(cell)
    return out


def _row_is_empty(row: dict[str, str]) -> bool:
    return all(not v for v in row.values())


def _normalise_rows(
    rows: list[list[Any]], lookup: dict[str, str],
) -> ParseResult:
    """Common path for any parser that has produced a list-of-rows
    from its wire format.  Pulls the header row, maps each
    subsequent row to an internal-field dict, drops blanks,
    returns the parse manifest."""
    if not rows:
        return ParseResult()
    raw_headers = [_stringy(h) for h in rows[0]]
    normalised = [_normalise_header(h) for h in raw_headers]
    mapped = sorted({h for h in normalised if h in lookup})
    unmapped = [
        raw for raw, norm in zip(raw_headers, normalised)
        if norm not in lookup and norm
    ]

    items: list[dict[str, str]] = []
    for row in rows[1:]:
        normalised_row = _row_to_normalised(raw_headers, row, lookup)
        if _row_is_empty(normalised_row):
            continue
        # Require an item name to keep a row.  An export without
        # name is essentially a placeholder cell.
        if not normalised_row.get("item_name"):
            continue
        items.append(normalised_row)
        if len(items) >= MAX_IMPORT_ROWS:
            break

    return ParseResult(
        items=items,
        unmapped_headers=unmapped,
        mapped_headers=mapped,
        total_rows=max(len(rows) - 1, 0),
    )


# ── Encircle parser (#1 of N) ───────────────────────────────────────
#
# Encircle's consumer "Detailed Spreadsheet" + the pro-tier variant
# share the same column conventions (the pro tier adds depreciation
# / ACV / tax / box columns).  Header strings drift slightly
# between exports — the lookup below carries the variants observed
# in the wild from HomeBox discussion #1065, the NesVentory repo,
# and Encircle's own help-center articles.


ENCIRCLE_HEADER_MAP: dict[str, set[str]] = {
    "item_name": {"name", "item", "item name", "description", "title"},
    "brand": {"manufacturer", "brand", "make"},
    "model_number": {
        "model number", "model #", "model", "model no", "model no.",
    },
    "serial_number": {
        "serial number", "serial #", "serial", "sn", "serial no",
    },
    "quantity": {"quantity", "qty", "count"},
    "purchase_vendor": {
        "purchase vendor", "purchased from", "vendor", "store",
        "retailer",
    },
    "purchase_date": {"purchase date", "date purchased", "date"},
    "purchase_price": {"purchase price", "price", "cost", "paid"},
    "estimated_value": {
        "estimated value", "value", "replacement cost", "current value",
    },
    "warranty_duration": {
        "warranty duration", "warranty", "warranty period",
    },
    "extended_warranty_policy": {
        "extended warranty policy", "warranty policy",
    },
    "extended_warranty_phone": {"extended warranty phone"},
    "notes": {"notes", "comments", "memo"},
    "room": {"room", "location", "structure"},
    "box": {"box"},
    "disposition": {"disposition"},
    "item_age": {"item age", "age"},
    "pre_loss_condition": {"pre-loss condition", "condition"},
    "upc": {"upc", "barcode"},
}


_ENCIRCLE_LOOKUP: dict[str, str] = {
    variant: internal
    for internal, variants in ENCIRCLE_HEADER_MAP.items()
    for variant in variants
}


def parse_encircle_csv(file_bytes: bytes) -> ParseResult:
    """Parse an Encircle CSV upload (the "Save As CSV" variant of
    the web app's Detailed Spreadsheet export).  Decodes as UTF-8
    with a BOM-tolerant fallback; uses ``csv.reader`` so embedded
    newlines in Notes survive intact."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return _normalise_rows(rows, _ENCIRCLE_LOOKUP)


def parse_encircle_xlsx(file_bytes: bytes) -> ParseResult:
    """Parse an Encircle XLSX upload (the web-app Detailed
    Spreadsheet export).  Reads only the first worksheet — the
    Encircle export is single-sheet.  V1 reads text only; photo
    extraction from ``xl/media/`` ships in a follow-up commit."""
    # openpyxl import is lazy so the rest of the module loads in
    # environments where openpyxl isn't installed (tests that
    # don't exercise the importer).
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return _normalise_rows(rows, _ENCIRCLE_LOOKUP)


# ── Parser registry ─────────────────────────────────────────────────


@dataclass
class Importer:
    """Per-source registration entry.

    ``label`` is the display name baked into the auto-created
    Location ("Imported from Encircle (…)").  ``parse_csv`` /
    ``parse_xlsx`` are the source-specific parsers (either may
    be None if the source doesn't ship that format)."""
    label: str
    parse_csv: Callable[[bytes], ParseResult] | None = None
    parse_xlsx: Callable[[bytes], ParseResult] | None = None


PARSERS: dict[str, Importer] = {
    "encircle": Importer(
        label="Encircle",
        parse_csv=parse_encircle_csv,
        parse_xlsx=parse_encircle_xlsx,
    ),
    # Future importers register here.  e.g.::
    #
    #   "sortly": Importer(
    #       label="Sortly",
    #       parse_csv=parse_sortly_csv,
    #       parse_xlsx=parse_sortly_xlsx,
    #   ),
    #   "homebox": Importer(
    #       label="HomeBox",
    #       parse_csv=parse_homebox_csv,
    #   ),
}


def parse(
    source: str, filename: str, file_bytes: bytes,
) -> ParseResult:
    """Dispatch to the registered parser for ``source``.  Picks
    the CSV vs XLSX branch off the filename's extension.

    Unknown source → ValueError so the route layer can return
    400; unknown extension → ValueError too so we don't try to
    feed PDF bytes to the CSV parser."""
    importer = PARSERS.get(source)
    if importer is None:
        raise ValueError(f"unknown import source {source!r}")
    if len(file_bytes) > MAX_IMPORT_BYTES:
        raise ValueError(
            f"upload too large: {len(file_bytes)} bytes "
            f"(max {MAX_IMPORT_BYTES})"
        )
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xlsm"):
        if importer.parse_xlsx is None:
            raise ValueError(
                f"{importer.label} importer doesn't support .{ext}"
            )
        result = importer.parse_xlsx(file_bytes)
    elif ext == "csv":
        if importer.parse_csv is None:
            raise ValueError(
                f"{importer.label} importer doesn't support .{ext}"
            )
        result = importer.parse_csv(file_bytes)
    else:
        raise ValueError(
            f"unsupported file extension {ext!r}; "
            "expected .csv or .xlsx"
        )
    result.source = source
    return result


# ── Notes formatter ─────────────────────────────────────────────────


def format_item_notes(parsed: dict[str, str]) -> str:
    """Pack the source-side metadata fields into a single notes blob.
    Stash's ``items.notes`` is free-text; the metadata external
    tools capture doesn't all have first-class column homes here
    yet, so we preserve it as a structured prose block that
    survives search + audit reads.

    Skip empty fields entirely so a sparse row doesn't produce
    notes full of empty sentinels."""
    lines: list[str] = []
    label_pairs = [
        ("Brand", parsed.get("brand")),
        ("Model", parsed.get("model_number")),
        ("Serial", parsed.get("serial_number")),
        ("UPC", parsed.get("upc")),
        ("Qty", parsed.get("quantity")),
        ("Condition", parsed.get("pre_loss_condition")),
        ("Age", parsed.get("item_age")),
    ]
    purchase_bits: list[str] = []
    if parsed.get("purchase_vendor"):
        purchase_bits.append(parsed["purchase_vendor"])
    if parsed.get("purchase_date"):
        purchase_bits.append(parsed["purchase_date"])
    if parsed.get("purchase_price"):
        purchase_bits.append(f"${parsed['purchase_price']}")
    if purchase_bits:
        label_pairs.append(("Purchase", " · ".join(purchase_bits)))
    if parsed.get("estimated_value"):
        label_pairs.append(("Est value", f"${parsed['estimated_value']}"))

    warranty_bits: list[str] = []
    if parsed.get("warranty_duration"):
        warranty_bits.append(parsed["warranty_duration"])
    if parsed.get("extended_warranty_policy"):
        warranty_bits.append(parsed["extended_warranty_policy"])
    if parsed.get("extended_warranty_phone"):
        warranty_bits.append(parsed["extended_warranty_phone"])
    if warranty_bits:
        label_pairs.append(("Warranty", " · ".join(warranty_bits)))

    for label, value in label_pairs:
        if value:
            lines.append(f"{label}: {value}")
    if parsed.get("notes"):
        lines.append("")
        lines.append(parsed["notes"])

    return "\n".join(lines).strip()


# ── Generic executor ────────────────────────────────────────────────


# Prefix for auto-created Locations.  ``execute_import`` appends
# the source label + a timestamp.  An Undo cascade-delete walks
# the Location → Floor → Rooms → Boxes → Items tree.
IMPORTED_LOCATION_PREFIX = "Imported from"


def execute_import(
    actor: Actor, items: list[dict[str, str]],
    *, source: str = "external",
    label_suffix: str | None = None,
) -> dict:
    """Bulk-create items in the actor's tenant.  Returns
    ``{location_id, location_name, item_count, room_count,
    error_count}``.

    Strategy (product-agnostic):

    * Get-or-create a per-import Location named ``Imported from
      <Source> (YYYY-MM-DD HH:MM)``.  Each import gets its own so
      :func:`undo_import` can cascade-delete cleanly.
    * Get-or-create one Floor under that location ("Imported").
    * For each unique ``room`` value in ``items``, get-or-create a
      Stash room under the floor.
    * Per-room, get-or-create a "Loose items" box via the existing
      loose-box helper (re-uses feature shipped for #39).
    * Create one ``items`` row per parsed dict with metadata
      packed into ``notes`` via :func:`format_item_notes`.

    Maintainer-only — like every write path in the tenant."""
    require_role(actor, "maintainer")
    if actor.tenant_id is None:
        raise ForbiddenError(f"{actor.email} has no active tenant")

    if not items:
        return {
            "location_id": None,
            "location_name": None,
            "item_count": 0,
            "room_count": 0,
            "error_count": 0,
        }

    importer = PARSERS.get(source)
    source_label = importer.label if importer else source
    stamp = (label_suffix
             or datetime.now().strftime("%Y-%m-%d %H:%M"))
    location_name = f"{IMPORTED_LOCATION_PREFIX} {source_label} ({stamp})"

    with db() as conn:
        cur = conn.execute(
            "INSERT INTO locations (name, tenant_id) VALUES (?, ?)",
            (location_name, actor.tenant_id),
        )
        location_id = int(cur.lastrowid)
        cur = conn.execute(
            "INSERT INTO floors (name, location_id, tenant_id) "
            "VALUES ('Imported', ?, ?)",
            (location_id, actor.tenant_id),
        )
        floor_id = int(cur.lastrowid)
        obs.write_audit(
            conn,
            tenant_id=actor.tenant_id,
            actor_email=actor.email,
            action="import.start",
            target_kind="location",
            target_id=location_id,
            metadata={
                "source": source,
                "location": location_name,
                "item_count": len(items),
            },
        )
        conn.commit()

    # Per-room cache so we only round-trip get-or-create once per
    # unique room name.  Keyed by lower-cased name.
    room_id_by_name: dict[str, int] = {}
    loose_box_id_by_room: dict[int, int] = {}

    item_count = 0
    error_count = 0

    for parsed in items:
        room_name = (parsed.get("room") or "Unsorted").strip() or "Unsorted"
        room_key = room_name.lower()
        if room_key not in room_id_by_name:
            with db() as conn:
                cur = conn.execute(
                    "INSERT INTO rooms "
                    "(name, floor_id, location_id, "
                    " x, y, w, h, tenant_id) "
                    "VALUES (?, ?, ?, 0, 0, 0, 0, ?)",
                    (room_name, floor_id, location_id, actor.tenant_id),
                )
                room_id_by_name[room_key] = int(cur.lastrowid)
                conn.commit()
        room_id = room_id_by_name[room_key]

        if room_id not in loose_box_id_by_room:
            loose_box_id_by_room[room_id] = (
                dao_boxes.get_or_create_loose_for_room(actor, room_id)
            )
        box_id = loose_box_id_by_room[room_id]

        name = parsed.get("item_name") or "Untitled"
        notes = format_item_notes(parsed)
        try:
            dao_items.create(
                actor, box_id,
                name=name, notes=notes, photo=None, source_photo=None,
            )
            item_count += 1
        except Exception as exc:  # noqa: BLE001
            _log.warning(
                "import.item_failed name=%r room=%r: %s",
                name, room_name, exc,
            )
            error_count += 1

    with db() as conn:
        obs.write_audit(
            conn,
            tenant_id=actor.tenant_id,
            actor_email=actor.email,
            action="import.complete",
            target_kind="location",
            target_id=location_id,
            metadata={
                "source": source,
                "item_count": item_count,
                "room_count": len(room_id_by_name),
                "error_count": error_count,
            },
        )
        conn.commit()

    _log.info(
        "import.complete tenant_id=%s source=%s location_id=%s "
        "items=%s rooms=%s errors=%s",
        actor.tenant_id, source, location_id, item_count,
        len(room_id_by_name), error_count,
    )

    return {
        "location_id": location_id,
        "location_name": location_name,
        "item_count": item_count,
        "room_count": len(room_id_by_name),
        "error_count": error_count,
    }


def undo_import(actor: Actor, location_id: int) -> dict:
    """Cascade-delete the auto-created import Location.  Tenant-
    scoped — refuses to delete a location the actor doesn't own
    or a location whose name doesn't carry the
    ``IMPORTED_LOCATION_PREFIX`` (defends against an operator
    accidentally typing a real location id into the undo URL)."""
    require_role(actor, "maintainer")
    if actor.tenant_id is None:
        raise ForbiddenError(f"{actor.email} has no active tenant")
    with db() as conn:
        row = conn.execute(
            "SELECT id, name FROM locations "
            "WHERE id = ? AND tenant_id = ?",
            (location_id, actor.tenant_id),
        ).fetchone()
        if row is None:
            raise NotFoundError(f"location {location_id}")
        if not str(row["name"]).startswith(IMPORTED_LOCATION_PREFIX):
            # Opaque to the operator that this location exists at
            # all — same rule the rest of the cross-tenant surface
            # uses for cross-scope probes.
            raise NotFoundError(f"location {location_id}")
        meta = conn.execute(
            "SELECT "
            "  (SELECT COUNT(*) FROM rooms WHERE location_id = ?) AS rooms, "
            "  (SELECT COUNT(*) FROM boxes b "
            "    JOIN rooms r ON r.id = b.room_id "
            "    WHERE r.location_id = ?) AS boxes, "
            "  (SELECT COUNT(*) FROM items i "
            "    JOIN boxes b ON b.id = i.box_id "
            "    JOIN rooms r ON r.id = b.room_id "
            "    WHERE r.location_id = ?) AS items",
            (location_id, location_id, location_id),
        ).fetchone()
        obs.write_audit(
            conn,
            tenant_id=actor.tenant_id,
            actor_email=actor.email,
            action="import.undo",
            target_kind="location",
            target_id=location_id,
            metadata=dict(meta),
        )
        conn.execute(
            "DELETE FROM locations WHERE id = ?", (location_id,),
        )
        conn.commit()
    _log.info(
        "import.undo tenant_id=%s location_id=%s removed",
        actor.tenant_id, location_id,
    )
    return dict(meta)
